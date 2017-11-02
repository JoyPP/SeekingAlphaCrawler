import urllib2
from bs4 import BeautifulSoup
import time
import cPickle
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
import os


def search_links(symbol, num_pages = 500):
    hdr = {'User-Agent': 'Mozilla/5.0'}

    link_list = list()
    # search all the articles about specific symbol and save it in link_list
    print 'search links for symbol ', symbol
    for i in range(1, num_pages):
        url = 'https://seekingalpha.com/symbol/'+ symbol + '/more_focus?page=' + str(i)
        req = urllib2.Request(url, headers=hdr)
        html = urllib2.urlopen(req).read()
        soup = BeautifulSoup(html)
        contents = soup.find_all("div", class_='\\"symbol_article\\"')
        if len(contents) == 0:
            break
        for c in contents:
            arti_link = c.findChild('a').get('href', None)
            full_link = "https://seekingalpha.com" + arti_link.replace('\\"','')
            link_list.append(full_link)
    print 'Searching links finished.'
    return link_list

def extract_summary(symbol, link_list):
    hdr = {'User-Agent': 'Mozilla/5.0'}
    content_list = list()
    for i in range(len(link_list)):
        url = link_list[i]
        print 'extract summary from page: ', url
        page_info = dict()
        try:
            req = urllib2.Request(url, headers=hdr)
            html = urllib2.urlopen(req).read()
            soup = BeautifulSoup(html)
        except:
            print 'the %d_th link %s error ' % (i, url)
            time.sleep(20)
            save_to_excel(symbol, content_list)
            content_list = list()
            try:
                req = urllib2.Request(url, headers=hdr)
                html = urllib2.urlopen(req).read()
                soup = BeautifulSoup(html)
            except:
                print 'still cannot open url ', url
                continue

        # get the main content
        main_content = soup.find_all("div", id = "content-rail")
        if len(main_content):
            main_content = main_content[0]
        else:
            continue

        # get the published time
        publishedTime = main_content.find_all("time", itemprop="datePublished")
        if len(publishedTime):
            publishedTime = publishedTime[0].get("content", None)
        else:
            continue
        if publishedTime.startswith('2014-02') or publishedTime.startswith('2013'):
            break

        # get title of the article
        title = main_content.find_all("h1", itemprop="headline")
        if len(title):
            title = title[0].text
        else:
            continue

        # get author homepage link
        author = main_content.find_all("a", class_="name-link")
        if len(author):
            author = author[0].get("href", None)
        else:
            continue

        # get summary
        summary = main_content.find_all("div", class_="a-sum")
        if len(summary):
            summary = summary[0]
            summary_content = list()
            paragraphs = summary.find_all("p")
            for p in paragraphs:
                summary_content.append(p.text)
        else:
            continue

        # save page information
        page_info['url'] = url
        page_info["publishedTime"] = publishedTime
        page_info["title"] = title
        page_info["author"] = author
        page_info["summary"] = summary_content
        content_list.append(page_info)
        time.sleep(3)
        if len(content_list) == 100:
            save_to_excel(symbol, content_list)
            content_list = list()
    return content_list

def save_links(symbol, link_list):
    filename = symbol + '_links.pkl'
    print 'save links into file ', filename
    with open(filename, 'w') as f:
        cPickle.dump(link_list, f)
    print 'saving links finished.'

def save_content(symbol, content_list):
    filename = symbol + '.pkl'
    print 'save content into file ', filename
    with open(filename, 'w') as f:
        cPickle.dump(content_list, f)
    print 'saving content finished.'


def save_to_excel(symbol, contents):
    filename = symbol + '.xlsx'
    print 'save content into ', filename
    if os.path.exists(filename):
        wb = load_workbook(filename=filename)
        ws = wb[symbol]
        row = ws.max_row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = symbol
        ws.cell(row=1, column=1, value='Title')
        ws.cell(row=1, column=2, value='PageLink')
        ws.cell(row=1, column=3, value='Time')
        ws.cell(row=1, column=4, value='Author')
        ws.cell(row=1, column=5, value='Summary')
        row = 1
    for i, c in zip(range(row + 1, row + 1 + len(contents)), contents):
        ws.cell(row=i, column=1, value=c['title'])
        ws.cell(row=i, column=2, value=c['url'])
        ws.cell(row=i, column=3, value=c['publishedTime'])
        ws.cell(row=i, column=4, value=c['author'])
        summary = c['summary']
        for j, s in zip(range(len(summary)), summary):
            ws.cell(row=i, column=5 + j, value=s)
    wb.save(filename)




if __name__ == '__main__':
    symbols = ['SYT', 'VALE']
    for symbol in symbols:
        print 'now get articles about ', symbol
        linkfile = symbol + '_links.pkl'
        if os.path.exists(linkfile):
            print 'load links from existing file ', linkfile
            with open(linkfile, 'r') as f:
                link_list = cPickle.load(f)
        else:
            link_list = search_links(symbol)
        print 'There are total ', len(link_list), ' links in the file.'
        save_links(symbol, link_list)
        content_list = extract_summary(symbol, link_list)
        save_to_excel(symbol, content_list)
