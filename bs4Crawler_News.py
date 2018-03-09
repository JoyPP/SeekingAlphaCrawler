import urllib2
from bs4 import BeautifulSoup
import time
from datetime import date, datetime
import cPickle
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
import os

def time_process(t):
    t = t.strip()
    add = 0
    if t.endswith("PM"):
        add = 12
    idx2 = t.find(" ")
    t = t[:idx2]
    hour, minute = t.split(":")
    hour = str(int(hour) + add)
    return ":".join([hour, minute, "00"])


def search_sector(sector):
    hdr = {'User-Agent': 'Mozilla/5.0'}
    sector_url = 'https://seekingalpha.com/market-news/' + sector + '?page='
    all_news = list()
    for page in range(792,1187):
        url = sector_url + str(page)
        while True:
            try:
                req = urllib2.Request(url, headers=hdr)
                html = urllib2.urlopen(req).read()
                soup = BeautifulSoup(html)
                break
            except:
                time.sleep(3)
        new_list = soup.find_all("ul", id='mc-list')[0].find_all("li")
        for item in new_list:
            one_new = dict()
            if item.has_attr("class"):
                if item["class"][0] == "date-title":
                    release_date = item.text
                    idx = release_date.find("-")
                    if idx >= 0:
                        release_date = release_date[idx+2:]
                    release_date = date.strftime(datetime.strptime(release_date, "%A, %B %d, %Y"),"%Y-%m-%d")
                elif item["class"][0] == "mc":
                    symbol = set([item.find("div", class_="media-left").text])
                    title = item.find("div", class_="title").text
                    details = item.find("div", class_="bullets")
                    content = details.text.strip()
                    links = details.find_all("a")
                    for l in links:
                        if l.has_attr("href"):
                            href = l["href"]
                            if href.startswith("https://seekingalpha.com/symbol/"):
                                symbol.add(href[href.rfind("/")+1:])
                    release_time = release_date + " " + time_process(item.find("span", class_="item-date").text)
                    one_new["symbol"] = ";".join(symbol)
                    one_new["title"] = title
                    one_new["content"] = content
                    one_new["time"] = release_time
                    all_news.append(one_new)
        if page % 10 == 0:
            print 'sector = %s, page = %d' % (sector, page)
            save_to_excel(sector, all_news)
            all_news = list()


def save_to_excel(sector, all_news, filename = 'News.xlsx'):
    print 'save content into ', filename
    if os.path.exists(filename):
        wb = load_workbook(filename=filename)
    else:
        wb = Workbook()
    if sector in wb.get_sheet_names():
        ws = wb[sector]
        row = ws.max_row
    else:
        ws = wb.create_sheet(sector)
        ws.cell(row=1, column=1, value='Title')
        ws.cell(row=1, column=2, value='Time')
        ws.cell(row=1, column=3, value='Symbol')
        ws.cell(row=1, column=4, value='Content')
        row = 1
    for i, c in zip(range(row + 1, row + 1 + len(all_news)), all_news):
        ws.cell(row=i, column=1, value=c['title'])
        ws.cell(row=i, column=2, value=c['time'])
        ws.cell(row=i, column=3, value=c['symbol'])
        ws.cell(row=i, column=4, value=c['content'])
    wb.save(filename)

def data_process(sectors, filename = 'News.xlsx'):
    print 'process data in', filename
    if os.path.exists(filename):
        wb = load_workbook(filename=filename)
    else:
        print 'File %s does not exist.' % filename
        return
    sector_info = dict()
    for sector in sectors:
        stocks = set([])
        text = ''
        ws = wb[sector]
        row = ws.max_row
        print 'sector %s has %d rows.' % (sector, row)
        for i in range(2, row+1):
            # modify symbols
            s = ws['C'+str(i)].value
            if s is not None and s.find(';') > 0:
                #print 'the original symbol =', s
                s = s.upper()
                l = set(s.split(';'))
                stocks.update(l)
                s = ";".join(l)
                #print 'after process =', s
                ws.cell(row=i, column=3, value=s)

            # save content to txt file
            title = ws['A'+str(i)].value
            content = ws['D'+str(i)].value
            if title is not None and content is not None:
                text += (title.encode('utf-8') + ' ' + content.encode('utf-8') + '\n')
            if i % 1000 == 0:
                with open('text.txt','a') as f:
                    f.write(text)
                    text = ''
                if i % 10000 == 0:
                    print 'finish %d row already.' % i
        with open('text.txt', 'a') as f:
            f.write(text)
        sector_info[sector] = stocks
    wb.save(filename)

    for k,v in sector_info.items():
        print 'sector %s has %d stocks.' % (k, len(v))

    # save the stock information in each sector
    with open('sector_stocks.pkl', 'w') as f:
        cPickle.dump(sector_info, f)



if __name__ == '__main__':
    sectors = ['tech', 'energy', 'healthcare', 'consumer', 'financials', 'commodities']
    for sector in sectors:
        search_sector(sector)
    #data_process(sectors)
    print 'finish'